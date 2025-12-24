import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------- PAGE CONFIG ----------------
st.set_page_config(
    page_title="Course Completion Dashboard",
    layout="wide"
)

st.title("📘 Course Completion Dashboard")

uploaded_file = st.file_uploader(
    "Upload Excel file (multiple sheets allowed)",
    type=["xlsx"]
)

# ---------------- MAIN LOGIC ----------------
if uploaded_file:

    xls = pd.ExcelFile(uploaded_file)
    combined_df = pd.DataFrame()

    # Column name possibilities (extended for your file)
    possible_name_cols = [
        "Employee Name", "Employee_name", "Name", "Name of the Official"
    ]

    possible_empno_cols = [
        "Employee No.", "Employee No", "Employee_id", "Employee ID", "Emp No"
    ]

    for sheet in xls.sheet_names:

        # Your file has headers in 2nd row → header=1
        df = pd.read_excel(uploaded_file, sheet_name=sheet, header=1)

        # Clean column names
        df.columns = df.columns.astype(str).str.strip()

        # Drop empty & unnamed columns
        df = df.dropna(axis=1, how="all")
        df = df[[c for c in df.columns if not c.lower().startswith("unnamed")]]

        # -------- Detect Employee Name column --------
        emp_name_col = None
        for col in possible_name_cols:
            if col in df.columns:
                emp_name_col = col
                break

        # -------- Detect Employee ID column --------
        emp_no_col = None
        for col in possible_empno_cols:
            if col in df.columns:
                emp_no_col = col
                break

        # Skip sheet if essential columns missing
        if emp_name_col is None or emp_no_col is None:
            continue

        # ---- IMPORTANT ----
        # Your file is ALREADY filtered (pending list)
        # So we DO NOT filter RMS TP or anything else
        df_valid = df.copy()

        # Add course name from sheet name
        df_valid["Course Name"] = sheet

        # Normalize column names
        df_valid = df_valid.rename(
            columns={
                emp_name_col: "Employee Name",
                emp_no_col: "Employee No."
            }
        )

        combined_df = pd.concat(
            [combined_df, df_valid],
            ignore_index=True
        )

    # ---------------- VALIDATION ----------------
    if combined_df.empty:
        st.error("No valid employee data found in the uploaded file.")
        st.stop()

    st.success("✅ Data extracted successfully")

    # Preview
    st.subheader("📄 Extracted Data Preview")
    st.dataframe(
        combined_df[
            ["Employee Name", "Employee No.", "Course Name"]
        ]
    )

    # ---------------- PIVOT TABLE ----------------
    st.subheader("📊 Course Completion Pivot")

    pivot_df = combined_df.pivot_table(
        index="Employee Name",
        columns="Course Name",
        values="Employee No.",
        aggfunc="count",
        fill_value=0
    )

    # Add totals
    pivot_df["Grand Total"] = pivot_df.sum(axis=1)
    pivot_df.loc["Grand Total"] = pivot_df.sum(numeric_only=True)

    st.dataframe(pivot_df)

    # ---------------- EXCEL EXPORT ----------------
    def export_pivot_to_excel(df: pd.DataFrame) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pivot Summary"

        for row in dataframe_to_rows(df.reset_index(), index=False, header=True):
            ws.append(row)

        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")

        # Header styling
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Bold Grand Total row
        for r in ws.iter_rows():
            if str(r[0].value).strip().lower() == "grand total":
                for c in r:
                    c.font = Font(bold=True)
                    c.alignment = center

        # Auto column width
        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    excel_bytes = export_pivot_to_excel(pivot_df)

    st.download_button(
        label="📥 Download Pivot Excel",
        data=excel_bytes.getvalue(),
        file_name="Course_Pivot_Summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
